#!/usr/bin/env python3
# coding=UTF-8
import openpyxl
import os

import sys

curDir = os.getcwd()  # 获取当前文件路径
rootDir = curDir[:curDir.find("excel\\") + len("excel\\")]  # 获取myProject，也就是项目的根路径
sys.path.append(rootDir)

import read


def summary_data(path):
    # pattern_column = '状态'
    # patter_value = '激活'
    # pattern_column = read.show_input_dialog('请选择过滤列名称', '')
    # patter_value = read.show_input_dialog('请选择过滤条件', '')

    two_dialog = read.show_two_dialog('', '', ['请选择过滤列名称', '请选择过滤条件'])
    pattern_column = two_dialog[0]
    patter_value = two_dialog[1]
    print(two_dialog)
    work_book = openpyxl.load_workbook(path, read_only=False)  #
    sheetnames = work_book.sheetnames
    print(sheetnames)
    # input_sheet_name = read.show_input_dialog('请选择需要汇总的表单，输入表单名或index序列号，用英文,号分割', sheetnames)
    # print(input_sheet_name)
    # summary_sheet_name = read.show_input_dialog('请选择要汇总到的表单，输入表单名或index序列号', sheetnames)
    # print(summary_sheet_name)

    two_dialog1 = read.show_two_dialog('请根据需要选择表单，输入表单名或index序列号，用英文,号分割', sheetnames, ['请选择需要汇总的表单', '请选择要汇总到的表单'])
    input_sheet_name = two_dialog1[0]
    summary_sheet_name = two_dialog1[1]
    print(input_sheet_name)
    print(summary_sheet_name)
    sheet_names = input_sheet_name.split(',')
    result_list = []
    # 过滤需要的表单信息
    for name in sheet_names:
        if name.isdigit():
            # print('sheet_name:' + sheetnames[int(name)])
            sheet = work_book.worksheets[int(name) - 1]
        else:
            # print('sheet_name:' + name)
            sheet = work_book[name]
        # 获取当前表单的所有行数
        rows = sheet.rows
        for index1, row in enumerate(rows):
            # print(row)
            line = [col.value for col in row]  # 取值
            if index1 == 0:
                # 第一行表头
                # print(line)
                # pattern_column = read.show_input_dialog('请选择过滤列名称', line)
                # patter_value = read.show_input_dialog('请选择过滤条件', '')
                pattern_column_index = line.index(pattern_column)
                # print(pattern_column_index)
                continue
            else:
                # print(line)
                if patter_value == line[pattern_column_index]:
                    result_list.insert(len(result_list), line)
                # break
    # print(result_list)
    # 读取要合并的表单
    if summary_sheet_name.isdigit():
        # print('sheet_name:' + sheetnames[int(summary_sheet_name) - 1])
        sheet1 = work_book.worksheets[int(summary_sheet_name) - 1]
    else:
        # print('sheet_name:' + summary_sheet_name)
        sheet1 = work_book[summary_sheet_name]
    max_row = sheet1.max_row
    print('最大行数%i;' % max_row)
    # 先删除原有的表单行数
    for i in range(0, max_row - 1):
        sheet1.delete_rows(max_row - i)
    # 写入
    for row_index in range(0, len(result_list)):
        for col_index in range(0, len(result_list[row_index])):
            sheet1.cell(row=row_index + 2, column=col_index + 1).value = result_list[row_index][col_index]
    work_book.save(path)


if __name__ == '__main__':
    # pattern_column = read.show_input_dialog('请选择过滤列名称', '')

    input_file_path = read.open_file_win('请选择需要合并的excel表', read.xlsx_file_types)
    for index, item in enumerate(input_file_path):
        print('读取第%i个文件:%s' % (index + 1, item))
        summary_data(item)
    # try:
    #     summary_data(item)
    # except:
    #     print('第%i个文件处理失败:%s' % (index + 1, item))
