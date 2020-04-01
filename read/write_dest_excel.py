#!/usr/bin/env python3
# coding=UTF-8
import os
import sys
import openpyxl

curDir = os.getcwd()  # 获取当前文件路径
rootDir = curDir[:curDir.find("excel\\") + len("excel\\")]  # 获取myProject，也就是项目的根路径
sys.path.append(rootDir)
import read

# 默认需要匹配的列标签
origin_pattern = '工号,岗位,身份证号,状态,店代码1,店代码2,店代码3,店代码4,店代码5'
dest_pattern = '工号,岗位,身份证号,状态,经销店代码1,经销店代码2,经销店代码3,经销店代码4,经销店代码5'


# 获取下标 pattern_list 需要匹配的列表，dest_list 被匹配的列表
def get_index_list(pattern_list, dest_list):
    list = []
    for index_, value in enumerate(pattern_list):
        # print('%i %s' % (index_, value))
        list.insert(index_, dest_list.index(value))
    return list


# 根据源文件工号唯一值修改目标文件相同工号的数据

def write_xlsx(read_path, write_path):
    # 读取xlsx目标文件
    work_book = openpyxl.load_workbook(write_path, read_only=False)  #
    sheetnames = work_book.sheetnames
    input_sheet_name = read.show_input_dialog('读取到如下表单，请选择需要修改的表单，输入表单名或index序列号，用英文,号分割:', sheetnames)
    # 先读取源文件
    if os.path.splitext(read_path)[1] == '.xls':
        read_data = read.read_excel_xls(read_path)
    else:
        read_data = read.read_excel_xlsx(read_path)
    # 获取源文件的表头
    origin_header = read_data['header']
    origin_rows = read_data['rows']
    print(origin_header)
    origin_index_list = get_index_list(origin_pattern.split(','), origin_header)
    # two_dialog = read.show_two_dialog('title', origin_header, ['匹配列标', ['同步的列数下标']])
    # print(two_dialog)
    # 需要操作的sheet集合
    sheet_names = input_sheet_name.split(',')
    # 循环打开sheet表单操作
    for name in sheet_names:
        if name.isdigit():
            sheet = work_book.worksheets[int(name) - 1]
        else:
            sheet = work_book[name]
        # 获取当前表单的所有行数
        rows = sheet.rows
        # columns = sheet.columns
        # 需要修改的当前sheet的表头，用来对比
        cur_sheet_header = []
        for index2, row in enumerate(rows):
            # print(row)
            line = [col.value for col in row]  # 取值
            if index2 == 0:
                # 第一行表头
                dest_index_list = get_index_list(dest_pattern.split(','), line)
                continue
            else:
                # 匹配行数修改内容
                dest_index = dest_index_list[0]
                key_value = line[dest_index]
                # print(key_value)
                # 获取元数据匹配的行
                origin_row_value = origin_rows[key_value]
                for index1, value in enumerate(origin_index_list):
                    value_ = origin_row_value[value]
                    col_num = dest_index_list[index1]
                    sheet.cell(row=index2 + 1, column=col_num + 1).value = value_
    work_book.save(write_path)


if __name__ == '__main__':
    read_file_path = read.open_single_file_win('请选择需要读取源目标excel表', read.xls_xlsx_file_types)
    print(read_file_path)
    write_file_path = read.open_multi_file_win('请选择需要合并的excel表', read.xlsx_file_types)
    for index, item in enumerate(write_file_path):
        print('读取第%i个文件:%s' % (index + 1, item))
        write_xlsx(read_file_path, item)
