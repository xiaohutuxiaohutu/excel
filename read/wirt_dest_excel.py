#!/usr/bin/env python3
# coding=UTF-8
# import xlrd
# import xlutils.copy
from xlrd import open_workbook
from xlwt import Workbook
from xlutils.copy import copy
import openpyxl
import os

file_path = 'C:\\Users\\23948\\Desktop\\excel\\人员列表汇理.xls'
dest_file_path = 'C:\\Users\\23948\\Desktop\\excel\\L1认证考试数据.xlsx'
dest_file_path_xlsx = 'C:\\Users\\23948\\Desktop\\excel\\L1认证考试数据.xlsx'
# 默认需要匹配的列标签
origin_pattern = '工号,岗位,身份证号,状态,店代码1,店代码2,店代码3,店代码4,店代码5'
dest_pattern = '工号,岗位,身份证号,状态,经销店代码1,经销店代码2,经销店代码3,经销店代码4,经销店代码5'
origin_pattern_index = []
dest_pattern_index = []

read_result_map = {}


# 读取xls
def read_excel_xls(path):
    read_excel_map = {}
    wb = open_workbook(path)
    sheets = wb.sheets()
    for sheet in sheets:
        # print(u"表单 %s 共 %d 行 %d 列" % (sheet.name, sheet.nrows, sheet.ncols))
        for row in range(0, sheet.nrows):
            values = sheet.row_values(row)
            if row == 0:
                read_result_map['header'] = values
                continue
            else:
                read_excel_map[values[0]] = values
    read_result_map['rows'] = read_excel_map
    return read_result_map


# 读取xlsx 返回类型{'header':'','rows':'{key:[],key1:[].....}'}
def read_excel_xlsx(path):
    read_excel_map = {}
    # 输入文件路径
    '''
    read_file_path = input("请输入读取文件路径:\n")
    print(read_file_path)
    '''
    work_book = openpyxl.load_workbook(path)  # 读取xlsx文件
    # names = data.get_sheet_names
    sheetnames = work_book.sheetnames
    print('读取到如下表单，请选择需要读取那几个表单，输入表单名或index序列号，用英文,号分割:')
    print(sheetnames)
    input_sheet_names = input()
    print(input_sheet_names)
    names_split = input_sheet_names.split(',')

    for name in names_split:
        if name.isdigit():
            index_ = int(name) - 1
            sheet = work_book.worksheets[int(name) - 1]

        else:
            sheet = work_book[name]
        rows = sheet.rows
        # print(type(rows))
        # columns = sheet.columns
        for index, row in enumerate(rows):
            # print(row)
            line = [col.value for col in row]  # 取值
            if index == 0:
                read_result_map['header'] = line
                continue
            else:
                # print(line)
                read_excel_map[line[0]] = line
    print(read_excel_map)
    read_result_map['rows'] = read_excel_map
    return read_result_map
    # excel_list.insert(index, line)
    # print(excel_list)
    # table = data.get_sheet_by_name(names[0])  # 获得指定名称的页
    # nrows = table.rows  # 获得行数 类型为迭代器
    # ncols = table.columns  # 获得列数 类型为迭代器
    # print(type(nrows))
    # for row in nrows:
    #     print(row)  # 包含了页名，cell，值
    #     line = [col.value for col in row]  # 取值
    #     print(line)
    # # 读取单元格
    # print(table.cell(1, 1).value)

    # read_excel_xlsx(dest_file_path_xlsx)


# 获取下标 pattern_list 需要匹配的列表，dest_list 被匹配的列表

def get_index_list(pattern_list, dest_list):
    list = []
    for index_, value in enumerate(pattern_list):
        # print('%i %s' % (index_, value))
        list.insert(index_, dest_list.index(value))
    return list


# 写入 xlsx文件
def write_xlsx(read_path, write_path):
    # 读取xlsx目标文件
    work_book = openpyxl.load_workbook(write_path, read_only=False)  #
    sheetnames = work_book.sheetnames
    print('读取到如下表单，请选择需要修改的表单，输入表单名或index序列号，用英文,号分割:')
    print(sheetnames)
    input_sheet_name = input()
    # 先读取源文件
    if os.path.splitext(read_path)[1] == '.xls':
        read_data = read_excel_xls(read_path)
    else:
        read_data = read_excel_xlsx(read_path)
    # 获取源文件的表头
    origin_header = read_data['header']
    origin_rows = read_data['rows']
    # print(origin_header)
    origin_index_list = get_index_list(origin_pattern.split(','), origin_header)
    # print(origin_index_list)
    # 需要操作的sheet集合
    sheet_names = input_sheet_name.split(',')
    # 循环打开sheet表单操作
    for name in sheet_names:
        if name.isdigit():
            sheet = work_book.worksheets[int(name) - 1]
        else:
            sheet = work_book[name]
        # 获取当前表单的所有行数
        rows = sheet.rows
        # columns = sheet.columns
        # 需要修改的当前sheet的表头，用来对比
        cur_sheet_header = []
        for index2, row in enumerate(rows):
            # print(row)
            line = [col.value for col in row]  # 取值

            if index2 == 0:
                # 第一行表头
                dest_index_list = get_index_list(dest_pattern.split(','), line)
                continue
            else:
                # print(line)
                # 匹配行数修改内容
                # 获取dest_index_list[0]
                dest_index = dest_index_list[0]
                key_value = line[dest_index]
                # print(key_value)
                # 获取元数据匹配的行
                origin_row_value = origin_rows[key_value]
                # print(origin_row_value)
                # print(origin_row_value)
                for index1, value in enumerate(origin_index_list):
                    value_ = origin_row_value[value]
                    col_num = dest_index_list[index1]
                    sheet.cell(row=index2 + 1, column=col_num + 1).value = value_
                    # print(sheet.cell(row=index2 + 1, column=col_num + 1).value)
                # break
    # work_book.close()
    work_book.save(write_path)


# 写入xls文件
def write_xls(read_path, write_path):
    print()


if __name__ == '__main__':
    read_file_path = input("请输入读取文件路径:\n")
    while read_file_path.strip() == '':
        read_file_path = input('读取文件路径不能为空，请重新输入：\n')
    while os.path.splitext(read_file_path)[1] not in ['.xls', '.xlsx']:
        read_file_path = input('文件类型不支持，目前只支持 xls xlsx，请重新输入：\n')

    write_file_path = input("请输入需要写入的文件路径：\n")
    while write_file_path.strip() == '':
        write_file_path = input('写入的文件路径不能为空，请重新输入：\n')
    # while os.path.splitext(write_file_path)[1] not in ['.xlsx']:
    #     write_file_path = input('写入的文件类型不支持，目前只支持xlsx，请重新输入：\n')

    input_org_match = input("请输入源文件需要匹配的列名，直接enter默认使用 工号,岗位,身份证号,状态,店代码1,店代码2,店代码3,店代码4,店代码5：\n")
    if input_org_match.strip() == '':
        input_org_match = '工号,岗位,身份证号,状态,店代码1,店代码2,店代码3,店代码4,店代码5'

    input_dest_match = input("请输入目标文件需要匹配的列名，直接enter默认使用 工号,岗位,身份证号,状态,经销店代码1,经销店代码2,经销店代码3,经销店代码4,经销店代码5：\n")
    if input_dest_match.strip() == '':
        input_dest_match = '工号,岗位,身份证号,状态,经销店代码1,经销店代码2,经销店代码3,经销店代码4,经销店代码5'
    while len(input_org_match.split(',')) != len(input_dest_match.split(',')):
        input_org_match = input('请重新输入源文件列名')
        if input_org_match.strip() == '':
            input_org_match = '工号,岗位,身份证号,状态,店代码1,店代码2,店代码3,店代码4,店代码5'
        input_dest_match = input("请重新输入目标文件需要匹配的列名，直接enter默认使用 工号,岗位,身份证号,状态,经销店代码1,经销店代码2,经销店代码3,经销店代码4,经销店代码5：\n")
        if input_dest_match.strip() == '':
            input_dest_match = '工号,岗位,身份证号,状态,经销店代码1,经销店代码2,经销店代码3,经销店代码4,经销店代码5'
    origin_pattern = input_org_match
    dest_pattern = input_dest_match
    # 校验输入的参数
    # 文件路径是否合法
    # 写入文件
    file_path = write_file_path.split(',')
    for index, item in enumerate(file_path):
        print('读取第%i个文件:%s' % (index + 1, item))
        write_xlsx(read_file_path, item)

        # while os.path.splitext(item)[1] not in ['.xlsx']:
        #     item = input('写入的文件类型不支持，目前只支持xlsx，请重新输入：\n')
        # try:
        #     write_xlsx(read_file_path, item)
        # except:
        #     print('第%i个文件处理失败:%s' % (index + 1, item))

# write_xlsx(read_file_path, write_file_path)
