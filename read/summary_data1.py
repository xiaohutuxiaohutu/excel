#!/usr/bin/env python3
# coding=UTF-8
import openpyxl
import os

import sys

curDir = os.getcwd()  # 获取当前文件路径
rootDir = curDir[:curDir.find("excel\\") + len("excel\\")]  # 获取myProject，也就是项目的根路径
sys.path.append(rootDir)

import read


def summary_data(path):
    work_book = openpyxl.load_workbook(path, read_only=False)  #
    sheetnames = work_book.sheetnames
    print(sheetnames)
    input_sheet_name = read.show_input_dialog('请选择需要读取的表名或位置', sheetnames)
    sheet_names = input_sheet_name.split(',')
    # 工号	姓名	岗位	状态	最高分数	店代码1	店代码2	店代码3	店代码4	店代码5
    # 过滤需要的表单信息
    sum_num = 0
    sheet_header = []
    for name in sheet_names:
        if name.isdigit():
            sheet_name = sheetnames[int(name)]
            sheet = work_book.worksheets[int(name) - 1]
        else:
            sheet_name = name
            sheet = work_book[name]
        # 获取当前表单的所有行数
        rows = sheet.rows
        max_row = sheet.max_row
        sum_num = sum_num + max_row
        print('%s表单处理前行数%i' % (sheet_name, max_row))
        # 最后结果集 {dian_dai_ma:[row]}
        all_map = {}
        for index1, row in enumerate(rows):
            # print(row)
            line = [col.value for col in row]  # 取值
            if index1 == 0:
                sheet_header = line
                # 第一行表头
                # print(line)
                continue
            else:
                dian_dai_ma_index = [5, 6, 7, 8, 9]
                # 当前行的最高分
                cur_max_score = line[4]
                for index2, value1 in enumerate(dian_dai_ma_index):
                    dian_dai_ma = line[value1]
                    if dian_dai_ma == '':
                        continue
                    # 店代码存在
                    elif dian_dai_ma in all_map.keys():
                        # 只保存通过或未通过的数据，最多有1条数据 通过或者未通过
                        dian_dai_ma_row = all_map[dian_dai_ma]
                        # print(dian_dai_ma_row)
                        max_score = dian_dai_ma_row[4]
                        # 当前行与已经保存的都是数字，说明是通过的，交换最高分
                        if max_score.isdigit() and cur_max_score.isdigit():
                            if float(cur_max_score) > float(max_score):
                                all_map[dian_dai_ma] = line
                            continue
                        # 都不是数字 忽略
                        elif not max_score.isdigit() and not cur_max_score.isdigit():
                            continue
                        # 已存在一个是数字，当前行是字符串，忽略
                        elif max_score.isdigit() and not cur_max_score.isdigit():
                            continue
                        # 已存在是字符串，当前行是数字，替换当前行
                        elif not max_score.isdigit() and cur_max_score.isdigit():
                            all_map[dian_dai_ma] = line
                    # 店代码不存在，直接插入
                    else:
                        all_map[dian_dai_ma] = line
    # 合并数据
    result_list1 = []
    for value2 in all_map.values():
        result_list1.insert(len(result_list1), value2)
    # 去重
    news_lists = []
    for id in result_list1:
        if id not in news_lists:
            news_lists.append(id)
    print('所有表单处理完毕 前 行数：%i' % sum_num)
    print('所有表单处理完毕 后 行数：%i' % len(news_lists))
    print('表头', end=':')
    print(sheet_header)
    news_lists.insert(0, sheet_header)
    create_sheet = work_book.create_sheet('删除后结果1', len(sheetnames))
    read.insert_sheet(news_lists, create_sheet, 1)
    work_book.save(path)


# 将指定表单内 每个店的数据只留一条（通过或删除），其他删除
if __name__ == '__main__':
    input_file_path = read.open_multi_file_win('请选择需要合并的excel表', read.xlsx_file_types)
    for index, item in enumerate(input_file_path):
        print('读取第%i个文件:%s' % (index + 1, item))
        summary_data(item)
    # try:
    #     summary_data(item)
    # except:
    #     print('第%i个文件处理失败:%s' % (index + 1, item))
