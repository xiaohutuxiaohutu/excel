# import xlrd
# import xlutils.copy
from xlrd import open_workbook
# from xlwt import Workbook
# from xlutils.copy import copy
import openpyxl
import tkinter as tk
from tkinter import filedialog
import tkinter.messagebox
import os

# 默认需要匹配的列标签
origin_pattern = '工号,岗位,身份证号,状态,店代码1,店代码2,店代码3,店代码4,店代码5'
dest_pattern = '工号,岗位,身份证号,状态,经销店代码1,经销店代码2,经销店代码3,经销店代码4,经销店代码5'
# 源文件需要读取的表头标签对应的列数集合
# origin_pattern_index = []
# 目标文件表头对应的列数集合
# dest_pattern_index = []


xls_file_types = [('excel文件', '.xls')]
xlsx_file_types = [('excel文件', '.xlsx')]
xls_xlsx_file_types = [('excel文件', '.xls'), ('excel文件', '.xlsx')]


# 打开文件多选窗口选择窗口
def open_multi_file_win(title, file_type):
    root = tk.Tk()
    root.withdraw()
    answer = filedialog.askopenfilenames(parent=root,
                                         initialdir=os.getcwd(),
                                         title=title,
                                         filetypes=file_type)
    if answer:
        root.destroy()
        return answer
    else:
        tkinter.messagebox.showinfo('提示', '没有选择文件，请重新选择')
        open_multi_file_win(title)


def open_single_file_win(title, file_type):
    root = tk.Tk()
    root.withdraw()
    answer = filedialog.askopenfilename(parent=root,
                                        initialdir=os.getcwd(),
                                        title=title,
                                        filetypes=file_type)
    if answer:
        root.destroy()
        return answer
    else:
        tkinter.messagebox.showinfo('提示', '没有选择文件，请重新选择')
        open_single_file_win(title)


# 单个输入弹框
def show_input_dialog(title, message):
    def return_callback(event):
        # print('quit...')
        root.quit()

    def close_callback():
        tk.messagebox.showinfo('message', 'no click...')

    root = tk.Tk(className=title)
    root.wm_attributes('-topmost', 1)
    screenwidth, screenheight = root.maxsize()
    width = 800
    height = 200
    size = '%dx%d+%d+%d' % (width, height, (screenwidth - width) / 2, (screenheight - height) / 2)
    root.geometry(size)
    root.resizable(0, 0)
    lable = tk.Label(root, height=2)
    lable['text'] = message
    lable.pack()
    entry = tk.Entry(root)
    entry.bind('<Return>', return_callback)
    entry.pack()
    entry.focus_set()
    root.protocol("WM_DELETE_WINDOW", close_callback)
    root.mainloop()
    str = entry.get()
    root.destroy()
    return str


# 读取xls所有表单的值 返回类型{'header':'','rows':'{key:[],key1:[].....}'}
def read_excel_xls(path):
    read_result_map = {}
    read_excel_map = {}
    wb = open_workbook(path)
    sheets = wb.sheets()
    for sheet in sheets:
        print(u"表单 %s 共 %d 行 %d 列" % (sheet.name, sheet.nrows, sheet.ncols))
        for row in range(0, sheet.nrows):
            values = sheet.row_values(row)
            if row == 0:
                read_result_map['header'] = values
                continue
            else:
                read_excel_map[values[0]] = values
    read_result_map['rows'] = read_excel_map
    return read_result_map


# 读取xlsx 返回类型{'header':'','rows':'{key:[],key1:[].....}'}
def read_excel_xlsx(path):
    read_excel_map = {}
    read_result_map = {}
    work_book = openpyxl.load_workbook(path)  # 读取xlsx文件
    # names = data.get_sheet_names
    sheetnames = work_book.sheetnames
    print('读取到如下表单，请选择需要读取那几个表单，输入表单名或index序列号，用英文,号分割:')
    print(sheetnames)
    input_sheet_names = input()
    print(input_sheet_names)
    names_split = input_sheet_names.split(',')

    for name in names_split:
        if name.isdigit():
            index_ = int(name) - 1
            sheet = work_book.worksheets[int(name) - 1]

        else:
            sheet = work_book[name]
        rows = sheet.rows
        # print(type(rows))
        # columns = sheet.columns
        for index, row in enumerate(rows):
            # print(row)
            line = [col.value for col in row]  # 取值
            if index == 0:
                read_result_map['header'] = line
                continue
            else:
                # print(line)
                read_excel_map[line[0]] = line
    # print(read_excel_map)
    read_result_map['rows'] = read_excel_map
    return read_result_map


# 获取下标 pattern_list 需要匹配的列表，dest_list 被匹配的列表
def get_index_list(pattern_list, dest_list):
    list = []
    for index_, value in enumerate(pattern_list):
        # print('%i %s' % (index_, value))
        list.insert(index_, dest_list.index(value))
    return list


# 根据源文件工号唯一值修改目标文件相同工号的数据 xlsx文件
def write_xlsx(read_path, write_path):
    # 读取xlsx目标文件
    work_book = openpyxl.load_workbook(write_path, read_only=False)  #
    sheetnames = work_book.sheetnames
    print('读取到如下表单，请选择需要修改的表单，输入表单名或index序列号，用英文,号分割:')
    print(sheetnames)
    input_sheet_name = input()
    # 先读取源文件
    if os.path.splitext(read_path)[1] == '.xls':
        read_data = read_excel_xls(read_path)
    else:
        read_data = read_excel_xlsx(read_path)
    # 获取源文件的表头
    origin_header = read_data['header']
    origin_rows = read_data['rows']
    # print(origin_header)
    origin_index_list = get_index_list(origin_pattern.split(','), origin_header)
    # print(origin_index_list)
    # 需要操作的sheet集合
    sheet_names = input_sheet_name.split(',')
    # 循环打开sheet表单操作
    for name in sheet_names:
        if name.isdigit():
            sheet = work_book.worksheets[int(name) - 1]
        else:
            sheet = work_book[name]
        # 获取当前表单的所有行数
        rows = sheet.rows
        # columns = sheet.columns
        # 需要修改的当前sheet的表头，用来对比
        cur_sheet_header = []
        for index2, row in enumerate(rows):
            # print(row)
            line = [col.value for col in row]  # 取值

            if index2 == 0:
                # 第一行表头
                dest_index_list = get_index_list(dest_pattern.split(','), line)
                continue
            else:
                # print(line)
                # 匹配行数修改内容
                # 获取dest_index_list[0]
                dest_index = dest_index_list[0]
                key_value = line[dest_index]
                # print(key_value)
                # 获取元数据匹配的行
                origin_row_value = origin_rows[key_value]
                # print(origin_row_value)
                # print(origin_row_value)
                for index1, value in enumerate(origin_index_list):
                    value_ = origin_row_value[value]
                    col_num = dest_index_list[index1]
                    sheet.cell(row=index2 + 1, column=col_num + 1).value = value_
                    # print(sheet.cell(row=index2 + 1, column=col_num + 1).value)
                # break
    # work_book.close()
    work_book.save(write_path)


# 写入xls文件
def write_xls(read_path, write_path):
    print()


def show_two_dialog(title, message, labels):
    master = tk.Tk(className=title)
    master.wm_attributes('-topmost', 1)
    screenwidth, screenheight = master.maxsize()
    width = 800
    height = 200
    size = '%dx%d+%d+%d' % (width, height, (screenwidth - width) / 2, (screenheight - height) / 2)
    master.geometry(size)
    master.resizable(0, 0)
    frame1 = tk.Frame(master)
    frame1.grid(row=0, column=0, sticky='w')
    tk.Label(frame1, text=message, justify='left').grid(row=0)

    frame2 = tk.Frame(master)
    frame2.grid(row=1, column=0, sticky='w')
    tk.Label(frame2, text=labels[0]).grid(row=1)
    e1 = tk.Entry(frame2)
    e1.grid(row=1, column=1, padx=10, pady=5)

    frame3 = tk.Frame(master)
    frame3.grid(row=2, column=0, sticky='w')
    tk.Label(frame3, text=labels[1]).grid(row=2)
    e2 = tk.Entry(frame3)
    e2.grid(row=2, column=1, padx=10, pady=5)

    frame4 = tk.Frame(master)
    frame4.grid(row=3, column=0, sticky='e')
    tk.Button(frame4, text="确定", width=10, command=master.quit).grid(row=3, column=1, sticky="e", padx=10, pady=5)
    master.mainloop()
    get1 = e1.get()
    get2 = e2.get()
    master.destroy()
    return [get1, get2]


def test_dialog(title, message, labels):
    # master = tk.Tk()
    master = tk.Tk(className=title)
    master.wm_attributes('-topmost', 1)
    screenwidth, screenheight = master.maxsize()
    width = 800
    height = 200
    size = '%dx%d+%d+%d' % (width, height, (screenwidth - width) / 2, (screenheight - height) / 2)
    master.geometry(size)
    master.resizable(0, 0)

    tk.Label(master, text=message).grid(row=0)
    label_values = []
    label_values1 = []
    for index1, value in enumerate(labels):
        cb1 = tk.Checkbutton('', text=value).grid(row=index1 + 2)
        print(cb1.get())
        tk.Label(master, text=value + "：").grid(row=index1 + 1)
        # e = tk.StringVar()

        label_values.insert(index1, tk.StringVar)
        # e = tk.Entry(master, textvariable=label_values[index1]).grid(row=index1 + 1, column=1, padx=10, pady=5)
        label_values1.insert(index1,
                             tk.Entry(master, textvariable=label_values[index1]).grid(row=index1 + 1, column=1, padx=10,
                                                                                      pady=5))
        # e = tk.StringVar()
        # entry = tk.Entry(master, textvariable=e).grid(row=index1 + 1, column=1, padx=10, pady=5)
        # entry.grid(row=index1 + 1, column=1, padx=10, pady=5)
        # entry.set('input your text here')
        # entry.pack()

        # e = tk.Entry(master)
        # e.grid(row=index1 + 1, column=1, padx=10, pady=5)
        # tk.Label(master, text="作者：").grid(row=2)

        # e1 = tk.Entry(master)
        # e2 = tk.Entry(master)
        # e1.grid(row=1, column=1, padx=10, pady=5)
        # e2.grid(row=2, column=1, padx=10, pady=5)

    def show():
        # print("作品：《%s》" % e1.get())
        # print("作者：%s" % e2.get())
        # e1.delete(0, "end")
        # e2.delete(0, "end")
        master.quit()
        # master.quit

    # tk.Button(master, text="获取信息", width=10, command=show).grid(row=3, column=0, sticky="w", padx=10, pady=5)
    # tk.Button(master, text="退出", width=10, command=show).grid(row=3, column=1, sticky="e", padx=10, pady=5)
    tk.Button(master, text="确定", width=10, command=master.quit).grid(row=len(labels) + 1, column=1, sticky="e", padx=10,
                                                                     pady=5)
    master.mainloop()
    result = {}
    # for index2, value in enumerate(labels):
    #     index3 = tk.Entry().index(index2 + 1)
    #
    #     result[value] = index3
    # value1 = e1.get()
    # value2 = e2.get()
    master.destroy()
    result_value = []
    for index5, vlaue1 in enumerate(label_values1):
        print(vlaue1)
    # for index4, value in enumerate(label_values):
    #     print(value)
    # print(tk.Entry(value).get())
    # result_value.insert(index4, tk.Entry().get())

    return label_values


# 插入数据到表单
def insert_sheet(data_list, sheet, start_row_index):
    for row_index in range(0, len(data_list)):
        for col_index in range(0, len(data_list[row_index])):
            sheet.cell(row=row_index + start_row_index, column=col_index + 1).value = data_list[row_index][col_index]


if __name__ == '__main__':
    print()
    str = '123ab4dgd'
    print(str.upper())
    str1 = ''
    print(str1)
    # dialog = test_dialog('title', 'sheet_name1,sheet_name2,sheetName3,sheent_name4', ['待合并表单名字', '需要合并到的表单名字', '待合并表单过滤列', '待合并表单过滤条件'])
    # dialog = show_two_dialog('title', 'sheet_name1,sheet_name2,sheetName3,sheent_name4', ['待合并表单名字', '需要合并到的表单名字', '待合并表单过滤列', '待合并表单过滤条件'])
    # dialog = show_two_dialog('请根据需要选择表单，输入表单名或index序列号，用英文,号分割', ['LI认证考试（20180816）', 'LI认证考试（20180126）', 'L1认证考试（20150128）', '黎经理专用版（截止20200225现存激活FAFC数据）'],

    # for i,value in enumerate(dialog):
    #   get = value().get()
    #   print(get)
    # print(dialog)
    # write_xlsx(dest_file_path_xlsx)


def filter_excel(workbook, column_name=0, by_name='Sheet0'):
    """

    :param workbook:
    :param column_name:
    :param by_name: 对应的Sheet页
    :return:
    """
    table = workbook.sheet_by_name(by_name)  # 获得表格
    total_rows = table.nrows  # 拿到总共行数
    columns = table.row_values(column_name)  # 某一行数据 ['姓名', '用户名', '联系方式', '密码']
    excel_list = []
    for one_row in range(1, total_rows):  # 也就是从Excel第二行开始，第一行表头不算

        row = table.row_values(one_row)
        # if row:
        #     row_object = {}
        #     for i in range(0, len(columns)):
        #         key = table_header[columns[i]]
        #         row_object[key] = row[i]  # 表头与数据对应
        #
        #     excel_list.append(row_object)

    return excel_list
