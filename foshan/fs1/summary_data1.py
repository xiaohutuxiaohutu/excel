#!/usr/bin/env python3
# coding=UTF-8
import openpyxl
import os

import sys

curDir = os.getcwd()  # 获取当前文件路径
rootDir = curDir[:curDir.find("excel\\") + len("excel\\")]  # 获取myProject，也就是项目的根路径
sys.path.append(rootDir)

import read

# 表单一需要读取的表头列名
sheet1_title = ['*实施事项名称', '*实施事项编码', '是否分情形(是|否)']
sheet2_title = ['*实施事项名称', '*实施事项编码', '*是否情形材料(是|否)', '情形名称(当情形材料时必填)', '*材料名称',
                '材料编号', '材料性质', '材料分类', '*是否纸质必需(是|否)', '*是否电子必需(是|否)', '*是否支持容缺(是|否)',
                '容缺情形', '*是否批文批复(是|否)', '*是否需要签章(是|否)', '原件数量', '复印件数量'
                ]
# 模板文件表头
write_sheet_title = ['*实施事项名称',
                     '*实施事项编码',
                     '*是否情形材料(是|否)',
                     '情形名称(当情形材料时必填)',
                     '父级情形名称',
                     '*材料名称',
                     '材料编号',
                     '材料性质',
                     '材料分类',
                     '*是否纸质必需(是|否)',
                     '*是否电子必需(是|否)',
                     '*是否支持容缺(是|否)',
                     '容缺情形',
                     '*是否批文批复(是|否)',
                     '*是否需要签章(是|否)',
                     '原件数量',
                     '复印件数量'
                     ]


def get_sheet1_row_values(work_book, start_index, col_list):
    # work_book = openpyxl.load_workbook(path, read_only=False)
    # 第一个表单 从第三行开始读取 读取列数 第二列 实施事项名称，第三列 实施事项编码 第六列  是否分情形，
    # sheet_1_col_index = [1, 2, 5]
    sheet_1 = work_book.worksheets[0]
    sheet_1_rows = sheet_1.rows
    # 读取表单1
    sheet1_row_value = []
    for sheet_1_row_index, row in enumerate(sheet_1_rows):
        line = [col.value for col in row]  # 取值
        if sheet_1_row_index < start_index:
            continue
        else:
            temp_value = []
            for line_index in col_list:
                new_value = line[line_index]
                temp_value.insert(len(temp_value), new_value)
            sheet1_row_value.insert(len(sheet1_row_value), temp_value)
    return sheet1_row_value


# 获取第二个表单敏感找事项分组
def get_sheet2_row_values(work_book, start_index, col_list):
    # sheet_2_col_index = [1, 4, 5, 7, 9, 10, 11, 13, 14, 15, 16]
    sheet_2_col_index = [1, 4, 5, 7, 9, 10, 11, 12, 13, 14, 15, 16]
    sheet_2 = work_book.worksheets[1]
    sheet_2_rows = sheet_2.rows

    sheet2_row_maps = {}
    for sheet_2_row_index, row in enumerate(sheet_2_rows):
        # print(sheet_0_row_index)
        line = [col.value for col in row]  # 取值
        if sheet_2_row_index == 0:
            continue
        else:
            # 匹配第二列事项名称
            line_item_name = line[1]
            if line_item_name is None or line_item_name == 'None':
                continue
            # print('line_item_name:' + line_item_name)
            # print(line_item_name)
            temp_value = []
            # 获取指定列的值
            for line_index in sheet_2_col_index:
                new_value = line[line_index]
                temp_value.insert(len(temp_value), new_value)
            if sheet_2_row_index == 1:
                print(temp_value)
            map_key_value = sheet2_row_maps.get(line_item_name)
            if map_key_value is None:
                sheet2_row_maps[line_item_name] = [temp_value]
            else:
                map_key_value.append(temp_value)
    return sheet2_row_maps


# 读取材料
def read_mat_code():
    file_path = read.open_single_file_win('选择原材料文件', read.xls_xlsx_file_types)
    work_book = openpyxl.load_workbook(file_path, read_only=False)
    sheet = work_book.worksheets[0]
    rows = sheet.rows
    # 读取表单1
    sheet_row_values = {}
    for row_index, row in enumerate(rows):
        line = [col.value for col in row]  # 取值
        if row_index == 0:
            continue
        else:
            mat_code = line[1]
            mat_name = line[2]
            sheet_row_values[mat_name] = mat_code
    return sheet_row_values


# 写入模板文件
def write_template(work_book, line1_value, line2_value, mat_codes):
    if line2_value is None or line2_value == 'None':
        return
    sheet = work_book.worksheets[0]
    max_row = sheet.max_row + 1
    # ['*实施事项名称', '情形名称(当情形材料时必填)', '*材料名称', '材料性质', '*是否纸质必需(是|否)', '*是否电子必需(是|否)', '*是否支持容缺(是|否)', '容缺情形', '*是否批文批复(是|否)', '*是否需要签章(是|否)', '原件数量', '复印件数量']
    value_len = len(line2_value)
    for row_index in range(0, value_len):
        row_value = line2_value[row_index]
        # print(row_value)
        # 写入首列序号
        index = max_row - 2 + row_index
        sheet.cell(row=row_index + max_row, column=1).value = index
        # 写入当前行的前三列
        # 第二列 事项名称
        sheet.cell(row=row_index + max_row, column=2).value = line1_value[0]
        # 第三列 事项编码
        sheet.cell(row=row_index + max_row, column=3).value = line1_value[1]
        # 第四列 *是否情形材料(是|否)
        sheet.cell(row=row_index + max_row, column=4).value = line1_value[2]
        # for col_index1 in range(0, len(line1_value)):
        #     sheet.cell(row=row_index + max_row, column=col_index1 + 2).value = line1_value[col_index1]
        # 插入第5列 情形名称
        if line1_value[2] == '是':
            sheet.cell(row=row_index + max_row, column=5).value = '情形一：' + row_value[1]
        else:
            sheet.cell(row=row_index + max_row, column=5).value = row_value[1]
        # 第6列 父级情形名称
        sheet.cell(row=row_index + max_row, column=6).value = '请选择办理的情形？'

        sheet_2_col_index = [1, 4, 5, 7, 9, 10, 11, 12, 13, 14, 15, 16]
        # 第7列 材料名称  个人身份证明 申请人身份证明
        mat_name = '个人身份证明' if row_value[2] == '申请人身份证明' else row_value[2]
        sheet.cell(row=row_index + max_row, column=7).value = mat_name
        # 8列 材料编号
        sheet.cell(row=row_index + max_row, column=8).value = mat_codes.get(mat_name)

        # 9列 材料性质
        sheet.cell(row=row_index + max_row, column=9).value = row_value[3]
        # 10列 材料分类
        sheet.cell(row=row_index + max_row, column=10).value = '申报材料'
        # 11   是否纸质必需(是|否)
        # sheet.cell(row=row_index + max_row, column=11).value = row_value[4]
        sheet.cell(row=row_index + max_row, column=11).value = '否'

        # 12 *是否电子必需(是|否)
        # sheet.cell(row=row_index + max_row, column=12).value = row_value[5]
        sheet.cell(row=row_index + max_row, column=12).value = '是'
        rong_que = '否' if (row_value[6] is None or row_value[6] == 'None') else row_value[6]
        # 13 * *是否支持容缺(是|否) ==None 否
        sheet.cell(row=row_index + max_row, column=13).value = rong_que

        # 14 容缺情形 ==None 忽略
        if row_value[7] is not None and row_value[7] != 'None':
            sheet.cell(row=row_index + max_row, column=14).value = row_value[7]
        # 15*是否批文批复(是|否) ==None 否
        pi_wen = '否' if (row_value[8] is None or row_value[8] == 'None') else row_value[8]
        sheet.cell(row=row_index + max_row, column=15).value = pi_wen

        # 16 *是否需要签章  ==None 是
        qian_zhang = '是' if (row_value[9] is None or row_value[9] == 'None') else row_value[9]
        sheet.cell(row=row_index + max_row, column=16).value = qian_zhang
        # 17 原件数量
        sheet.cell(row=row_index + max_row, column=17).value = row_value[10]
        # 18 复印件数量
        sheet.cell(row=row_index + max_row, column=18).value = row_value[11]


def summary_data():
    path = read.open_single_file_win('请选择需要处理的excel表', read.xls_xlsx_file_types)
    work_book = openpyxl.load_workbook(path, read_only=False)
    sheet1_row_values = get_sheet1_row_values(work_book, 1, [1, 2, 5])
    sheet2_map_values = get_sheet2_row_values(work_book, 1, [])
    # 读取材料编码和名称文件
    mat_codes = read_mat_code()
    # 读取模板文件
    dest_path = read.open_single_file_win('请选择需要写入的目标模板文件', read.xlsx_file_types)
    work_book = openpyxl.load_workbook(dest_path)
    for index3, value in enumerate(sheet1_row_values):
        if index3 == 0:
            continue
        else:
            item_name = value[0]
            # print('item_name:' + item_name)
            if str(item_name).startswith('禅城区'):
                new_item_name = str(item_name)[3:]
                # print('new_item_name:' + new_item_name)
                map_values = sheet2_map_values.get(new_item_name)
                if map_values is not None or map_values != 'None':
                    # print(map_values)
                    write_template(work_book, value, map_values, mat_codes)
        # print()

    work_book.save(dest_path)


# 将指定表单内 每个店的数据只留一条（通过或删除），其他删除
if __name__ == '__main__':
    # input_file_path = read.open_single_file_win('请选择需要处理的excel表', read.xls_xlsx_file_types)
    # des_file_path = read.open_single_file_win('请选择需要写入的目标模板文件', read.xlsx_file_types)
    summary_data()
